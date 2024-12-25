from ortools.linear_solver import pywraplp
import pandas as pd
import data_preprocess
from parse_course_time import TimeParser, TimeMatrix
from openpyxl.styles import Alignment


def matching_task(course_file_path, supervisor_file_path):

    # 加载并处理数据
    supervisor_course_data, task_data, fliter_data = data_preprocess.get_data(course_file_path, supervisor_file_path)

    # 解析课表信息
    time_parser_supervisor = TimeParser(supervisor_course_data, '课表信息')
    time_parser_task = TimeParser(task_data, '课表信息')
    all_schedule_list_supervisor = time_parser_supervisor.parse_schedules()
    all_schedule_list_task = time_parser_task.parse_schedules()

    # 创建时间矩阵用于冲突检测和时间接近度计算
    time_matrix = TimeMatrix(all_schedule_list_supervisor, all_schedule_list_task)

    num_supervisors = len(supervisor_course_data)
    num_tasks = len(task_data)
    
    # 创建线性规划模型
    solver = pywraplp.Solver.CreateSolver('SCIP')
    if not solver:
        print('无法创建求解器')
        return
    
    # 创建决策变量
    x = {}
    for i in range(num_supervisors):
        for j in range(num_tasks):
            x[i, j] = solver.BoolVar(f'x_{i}_{j}')
    
    # 添加硬约束条件: 每个任务只能被一个督导完成
    for j in range(num_tasks):
        solver.Add(solver.Sum([x[i, j] for i in range(num_supervisors)]) == 1)

    # 添加硬约束条件：均分所有任务
    tasks_per_supervisor = num_tasks // num_supervisors
    extra_tasks = num_tasks % num_supervisors

    for i in range(num_supervisors):
        if i < extra_tasks:
            # 如果有多余的任务，前 extra_tasks 个督导分配 tasks_per_supervisor + 1 个任务
            solver.Add(solver.Sum([x[i, j] for j in range(num_tasks)]) == tasks_per_supervisor + 1)
        else:
            # 其余的督导分配 tasks_per_supervisor 个任务
            solver.Add(solver.Sum([x[i, j] for j in range(num_tasks)]) == tasks_per_supervisor)

    # 添加硬约束条件：督导的课程时间不能与任务的时间冲突
    conflict_matrix, proximity_matrix = time_matrix.build_conflict_matrix()
    for i in range(num_supervisors):
        for j in range(num_tasks):
            if conflict_matrix[i][j]:
                solver.Add(x[i, j] == 0)
    
    # 添加软约束条件：督导的开课校区与任务的开课校区相同优先匹配
    campus_match = {}
    for i in range(num_supervisors):
        for j in range(num_tasks):
            supervisor_campuses = supervisor_course_data.loc[i, '开课校区'].split(',')
            task_campus = task_data.loc[j, '开课校区']
            campus_match[i, j] = 1 if task_campus in supervisor_campuses else 0

    # 添加软约束：听课时间与督导的课程时间越近的优先匹配
    proximity_score = {}
    for i in range(num_supervisors):
        for j in range(num_tasks):
            proximity_score[i, j] = proximity_matrix[i][j]
    
    # 添加目标函数：最大化匹配成功数
    solver.Maximize(solver.Sum([x[i, j] * (campus_match[i, j] + proximity_score[i, j]) for i in range(num_supervisors) for j in range(num_tasks)]))

    # 运行优化
    status = solver.Solve()
    
    #if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
    if status == pywraplp.Solver.OPTIMAL:
        print('找到最优解')
        assignment = []
        for i in range(num_supervisors):
            for j in range(num_tasks):
                if (i, j) in x and x[i, j].solution_value() > 0.5:
                    assignment.append((i, j))
        assignment_df = pd.DataFrame(assignment, columns=['督导', '听课任务'])
        # assignment_df.to_excel('./督导与听课任务匹配结果.xlsx', index=False)
        # print('匹配结果已保存到督导与听课任务匹配结果.xlsx')

        # 打印出每个督导分配的任务数量
        supervisor_task_count = [0] * num_supervisors
        for i, j in assignment:
            supervisor_task_count[i] += 1
        for i, count in enumerate(supervisor_task_count):
            print(f'督导 {i} 分配了 {count} 个任务')
        
        # output_data = pd.read_excel('./2024春开课任务_new_v4.xlsx')
        output_data = fliter_data
        output_data['督导姓名'] = ''
        for i, j in assignment:
            supervisor_name = supervisor_course_data.loc[i, '任课教师']
            course_name = task_data.loc[j, '课程名称']
            course_type = task_data.loc[j, '课程类型']
            course_desc = task_data.loc[j, '情况说明']
            
            mask = (
            (output_data['课程名称'] == course_name) &
            (output_data['课程类型'] == course_type) &
            (output_data['情况说明'] == course_desc)
            )
            
            output_data.loc[mask, '督导姓名'] = supervisor_name
        # 按课程名称、课程类型、情况说明排序、督导姓名排序
        output_data = output_data.sort_values(by=['课程名称', '课程类型', '情况说明', '督导姓名', '任课教师']).reset_index(drop=True)
        return output_data

    else:
        print('未找到可行解')

        return None


def save_to_excel(output_data, output_file):
    # 创建ExcelWriter对象
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入数据
        output_data.to_excel(writer, index=False, sheet_name='Sheet1')

        # 获取工作表
        worksheet = writer.sheets['Sheet1']

        # 全部居中对齐
        for row in worksheet.iter_rows(min_row=1, max_row=len(output_data) + 1, min_col=1, max_col=len(output_data.columns) + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 获取列索引
        condition_col_idx = output_data.columns.get_loc('情况说明') + 1  # Excel列从1开始
        supervisor_col_idx = output_data.columns.get_loc('督导姓名') + 1  # Excel列从1开始
        teacher_col_idx = output_data.columns.get_loc('任课教师') + 1  # Excel列从1开始
        
        # 初始化合并起始行
        start_row = 2  # 数据从第2行开始（第1行是表头）
        
        # 遍历每一行
        for row in range(3, len(output_data) + 2):  # +2是因为有表头
            # 获取当前行的关键字段值
            current_row_course = worksheet.cell(row=row, column=output_data.columns.get_loc('课程名称') + 1).value
            current_row_type = worksheet.cell(row=row, column=output_data.columns.get_loc('课程类型') + 1).value
            current_row_condition = worksheet.cell(row=row, column=output_data.columns.get_loc('情况说明') + 1).value
            current_row_supervisor = worksheet.cell(row=row, column=output_data.columns.get_loc('督导姓名') + 1).value
            current_row_teacher = worksheet.cell(row=row, column=teacher_col_idx).value
            
            # 获取上一行的关键字段值
            prev_row_course = worksheet.cell(row=row-1, column=output_data.columns.get_loc('课程名称') + 1).value
            prev_row_type = worksheet.cell(row=row-1, column=output_data.columns.get_loc('课程类型') + 1).value
            prev_row_condition = worksheet.cell(row=row-1, column=output_data.columns.get_loc('情况说明') + 1).value
            prev_row_supervisor = worksheet.cell(row=row-1, column=output_data.columns.get_loc('督导姓名') + 1).value
            prev_row_teacher = worksheet.cell(row=row-1, column=teacher_col_idx).value
            
            # 检查课程名称和类型是否改变
            values_changed = (
                current_row_course != prev_row_course or
                current_row_type != prev_row_type or
                current_row_condition != prev_row_condition or
                current_row_supervisor != prev_row_supervisor or
                (current_row_condition == '在该老师这门课程中任选其一' and current_row_teacher != prev_row_teacher) or
                row == len(output_data) + 1
            )
            
            # 如果值不同或到达最后一行，执行合并
            if values_changed:
                if row - start_row >= 1:  # 只有当有多行相同值时才合并
                    # 获取起始行的情况说明
                    start_row_condition = worksheet.cell(row=start_row, column=condition_col_idx).value
                    
                    # 只有当情况说明不是"必选"时才合并单元格
                    if start_row_condition != '必选':
                        # 合并单元格
                        worksheet.merge_cells(
                            start_row=start_row,
                            end_row=row - 1 if row != len(output_data) + 1 else row,
                            start_column=condition_col_idx,
                            end_column=condition_col_idx
                        )
                        # 合并单元格
                        worksheet.merge_cells(
                            start_row=start_row,
                            end_row=row - 1 if row != len(output_data) + 1 else row,
                            start_column=supervisor_col_idx,
                            end_column=supervisor_col_idx
                        )
                start_row = row
    
    print(f'匹配结果已保存到{output_file}')



if __name__ == "__main__":
    course_file_path = '../test-algorithm/2024春开课任务.xls'
    supervisor_file_path = '../test-algorithm/督导名单.xlsx'
    output_data = matching_task(course_file_path, supervisor_file_path)
    
    if output_data is not None:
        # 创建ExcelWriter对象
        output_file = '../test-algorithm/2024春开课任务_结果_合并.xlsx'
        save_to_excel(output_data, output_file)